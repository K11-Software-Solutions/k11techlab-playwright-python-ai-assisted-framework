# Renamed from data_reader_util.py

import sqlite3
import json
import csv
import openpyxl

def read_sql_data(db_path: str, query: str):
	"""
	Reads data from a SQL database using the provided query and returns a list of tuples.
	Example usage:
		data = read_sql_data('test.db', 'SELECT email, password FROM users')
	"""
	data = []
	try:
		conn = sqlite3.connect(db_path)
		cursor = conn.cursor()
		cursor.execute(query)
		data = cursor.fetchall()
		conn.close()
	except Exception as e:
		print(f"Error reading SQL database: {e}")
	return data

def read_json_data(file_path: str):
	data = []
	try:
		with open(file_path, "r", encoding="utf-8") as file:
			json_data = json.load(file)
			for record in json_data:
				data.append(tuple(record.values()))
	except Exception as e:
		print(f"Error reading JSON file: {e}")
	return data

def read_csv_data(file_path: str):
	data = []
	try:
		with open(file_path, newline='', encoding='utf-8') as file:
			reader = csv.DictReader(file)
			for row in reader:
				row.pop('validity', None)
				data.append(tuple(row.values()))
	except Exception as e:
		print(f"Error reading CSV file: {e}")
	return data

def read_excel_data(file_path: str, sheet_name: str = None):
	data = []
	try:
		workbook = openpyxl.load_workbook(file_path)
		sheet = workbook[sheet_name] if sheet_name else workbook.active
		headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
		validity_idx = None
		if 'validity' in headers:
			validity_idx = headers.index('validity')
		for row in sheet.iter_rows(min_row=2, values_only=True):
			if validity_idx is not None:
				row = tuple(val for i, val in enumerate(row) if i != validity_idx)
			data.append(row)
	except Exception as e:
		print(f"Error reading Excel file: {e}")
	return data
